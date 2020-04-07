import xlsxwriter

import openpyxl as openpyxl
import xlrd
from pandas.tests.io.excel.test_openpyxl import openpyxl
import pytest
from openpyxl import Workbook

import pandas as pd


def Identification(value, sheets):  # Save to the Excel Sheet
    wbkName = 'Data\SDS.xlsx'
    wbk = openpyxl.load_workbook(wbkName)
    sheet = wbk[sheets]
    sheet["D3"] = value[0]
    sheet["D5"] = value[1]
    sheet["D6"] = value[2]
    sheet["D7"] = value[3]
    sheet["D8"] = value[4]
    sheet["D9"] = value[5]
    sheet["D10"] = value[6]
    sheet["D12"] = value[7]
    sheet["D13"] = value[8]
    sheet["D15"] = value[9]
    sheet["D17"] = value[10]

    wbk.save(wbkName)
    wbk.close()


def Hazard(value, sheets):
    wbkName = 'Data\SDS.xlsx'
    wbk = openpyxl.load_workbook(wbkName)

    sheet = wbk[sheets]
    sheet["E3"] = value[0]
    sheet["E4"] = value[1]
    sheet["E6"] = value[2]
    sheet["E7"] = value[3]
    sheet["E9"] = value[4]
    sheet["E10"] = value[5]
    sheet["E11"] = value[6]
    sheet["E12"] = value[7]
    sheet["E14"] = value[8]

    wbk.save(wbkName)
    wbk.close()


def image_insert():
    # wb = openpyxl.Workbook("Data\SDS.xlsx")
    openpyxl_version = openpyxl.__version__
    workbook = xlsxwriter.Workbook("Data\SDS_Standard.xlsx")

    worksheet = workbook.add_worksheet()
    worksheet.set_column("B1:B5", 7)
    worksheet.set_default_row(45)
    image = ['Image15.png',
             'Image16.png']
    image_row = 0
    image_col = 0
    for img in image:
        worksheet.insert_image(image_row, image_col, img,
                               {'x_scale': 0.5, 'y_scale': 0.5, 'x_offset': 5, 'y_offset': 5, 'positioning': 1})
        image_row += 1

    worksheet.write(1, 1, "Label")
    workbook.close()
